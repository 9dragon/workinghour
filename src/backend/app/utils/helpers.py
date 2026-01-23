"""
工具函数
"""
import os
import json
from datetime import datetime, timedelta, date
import pandas as pd
from openpyxl import load_workbook

def generate_batch_no(prefix='IMP'):
    """生成批次号"""
    now = datetime.now()
    timestamp = now.strftime('%Y%m%d%H%M%S')
    import random
    random_num = random.randint(1000, 9999)
    return f'{prefix}_{timestamp}_{random_num}'

def validate_excel_file(file_path):
    """
    验证Excel文件

    返回: (is_valid, error_message)
    """
    try:
        # 检查文件扩展名
        if not file_path.endswith(('.xls', '.xlsx')):
            return False, '文件格式错误，仅支持.xlsx或.xls格式'

        # 检查文件大小（10MB限制）
        file_size = os.path.getsize(file_path)
        if file_size > 10 * 1024 * 1024:
            return False, '文件大小超过10MB限制'

        # 尝试读取文件
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # 检查是否有数据
        if ws.max_row < 2:
            return False, 'Excel文件没有数据'

        return True, None

    except Exception as e:
        return False, f'文件解析失败: {str(e)}'

def read_excel_data(file_path):
    """
    读取Excel数据，自动检测并兼容单行表头和双行表头格式

    支持的表头格式：
    1. 单行表头：第一行即为完整的列名（精简版）
    2. 双行表头：第二行才是完整的列名，第一行有合并单元格（原始版）

    返回: DataFrame or None
    """
    try:
        # 先读取前3行，检测表头格式
        df_sample = pd.read_excel(file_path, header=None, nrows=3)

        # 最关键的区分字段：项目交付-项目名称
        # 精简版第一行包含此字段
        # 原始版第一行列12-15是"当前项目交付工时"（重复），第二行才有完整字段名
        critical_field = '项目交付-项目名称'

        # 检测第一行是否包含关键字段
        first_row_str = str(df_sample.iloc[0].tolist())
        has_critical_in_first = critical_field in first_row_str

        # 如果第一行包含"项目交付-项目名称"，说明是精简版（单行表头）
        if has_critical_in_first:
            # 单行表头格式，直接读取
            df = pd.read_excel(file_path, header=0)
            df.columns = [str(col).strip() for col in df.columns]

            # 处理合并单元格导致的空值：使用前向填充
            for col in df.columns:
                if '部门' in col or col == '部门':
                    # 先将空字符串替换为NaN，然后进行前向填充
                    df[col] = df[col].replace('', pd.NA)
                    df[col] = df[col].fillna(method='ffill')

            return df

        # 检测第二行是否包含关键字段（原始版双行表头）
        second_row_str = str(df_sample.iloc[1].tolist())
        has_critical_in_second = critical_field in second_row_str

        if has_critical_in_second:
            # 双行表头格式
            # 使用openpyxl读取，避免pandas对合并单元格的处理问题
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            # 读取第二行作为列名
            headers = []
            for cell in ws[2]:  # 第二行
                headers.append(str(cell.value).strip() if cell.value else f'列{cell.column}')

            # 读取数据行（从第三行开始）
            data = []
            for row in ws.iter_rows(min_row=3, values_only=True):
                data.append(list(row))

            wb.close()

            # 创建DataFrame
            df = pd.DataFrame(data, columns=headers)

            # 处理合并单元格导致的空值：使用前向填充
            # 主要处理部门字段等可能合并的列
            for col in df.columns:
                if '部门' in col or col == '部门':
                    # 先将空字符串替换为NaN，然后进行前向填充
                    df[col] = df[col].replace('', pd.NA)
                    df[col] = df[col].fillna(method='ffill')

            return df

        # 如果都不匹配，尝试默认读取第一行
        df = pd.read_excel(file_path, header=0)
        df.columns = [str(col).strip() for col in df.columns]

        # 处理合并单元格导致的空值：使用前向填充
        for col in df.columns:
            if '部门' in col or col == '部门':
                # 先将空字符串替换为NaN，然后进行前向填充
                df[col] = df[col].replace('', pd.NA)
                df[col] = df[col].fillna(method='ffill')

        return df

    except Exception as e:
        print(f"读取Excel失败: {str(e)}")
        return None

def validate_work_hour_row(row, required_fields):
    """
    验证工时数据行

    返回: (is_valid, error_list)
    error_list: [{'field': 'field_name', 'error': 'error_message'}, ...]
    """
    errors = []

    # 检查必填字段
    for field in required_fields:
        if pd.isna(row.get(field, '')) or str(row.get(field, '')).strip() == '':
            errors.append({'field': field, 'error': f'{field}字段为空'})

    # 检查审批结果和状态
    approval_result = str(row.get('审批结果', '')).strip()
    approval_status = str(row.get('审批状态', '')).strip()

    # 审批结果必须是"通过"或"审批通过"
    if approval_result not in ['通过', '审批通过']:
        errors.append({'field': '审批结果', 'error': f"审批结果为'{approval_result}'，仅支持'通过'或'审批通过'"})

    # 审批状态必须是"已完成"或"已结束"
    if approval_status not in ['已完成', '已结束']:
        errors.append({'field': '审批状态', 'error': f"审批状态为'{approval_status}'，仅支持'已完成'或'已结束'"})

    # 检查工作时长
    try:
        work_val = row.get('项目交付-工作时长', 0)
        # 处理空字符串、None、NaN等情况
        if pd.isna(work_val) or str(work_val).strip() == '':
            work_hours = 0.0
        else:
            work_hours = float(str(work_val).strip())

        if work_hours < 0:
            errors.append({'field': '项目交付-工作时长', 'error': '工作时长不能为负数'})
        if work_hours > 168:
            errors.append({'field': '项目交付-工作时长', 'error': '工作时长超过168小时（一周最大时长）'})
    except Exception as e:
        errors.append({'field': '项目交付-工作时长', 'error': f'工作时长格式错误: {str(e)}'})

    # 检查加班时长
    try:
        overtime_val = row.get('项目交付-加班时长', 0)
        # 处理空字符串、None、NaN等情况
        if pd.isna(overtime_val) or str(overtime_val).strip() == '':
            overtime_hours = 0.0
        else:
            overtime_hours = float(str(overtime_val).strip())

        if overtime_hours < 0:
            errors.append({'field': '项目交付-加班时长', 'error': '加班时长不能为负数'})
    except Exception as e:
        errors.append({'field': '项目交付-加班时长', 'error': f'加班时长格式错误: {str(e)}'})

    if errors:
        return False, errors
    return True, []

def calculate_date_range(start_date, end_date, max_days=90):
    """
    计算日期范围，验证不超过max_days天

    返回: (is_valid, error_message, days_count)
    """
    try:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()

        if start > end:
            return False, "开始日期不能晚于结束日期", 0

        days_count = (end - start).days
        if days_count > max_days:
            return False, f"时间范围不能超过{max_days}天", days_count

        return True, None, days_count
    except ValueError as e:
        return False, "日期格式错误，应为YYYY-MM-DD", 0

def get_workdays_in_range(start_date, end_date, workdays, holidays=None):
    """
    获取指定时间范围内的工作日列表（考虑节假日）

    参数:
        start_date: 开始日期 (datetime.date)
        end_date: 结束日期 (datetime.date)
        workdays: 工作日列表 [1,2,3,4,5] (1=周一, 7=周日)
        holidays: 节假日列表 [datetime.date, ...] (可选)

    返回: 工作日日期列表
    """
    workdays_list = []
    current = start_date

    # 将节假日转换为集合以提高查找效率
    holiday_set = set(holidays) if holidays else set()

    while current <= end_date:
        # 检查是否是工作日（周一到周五）
        is_weekday_workday = current.weekday() + 1 in workdays

        # 检查是否是节假日
        is_holiday = current in holiday_set

        # 如果是工作日且不是节假日，则添加到列表
        if is_weekday_workday and not is_holiday:
            workdays_list.append(current)

        current += timedelta(days=1)

    return workdays_list

def calculate_workdays(start_date, end_date, workdays=[1, 2, 3, 4, 5], holidays=None):
    """
    计算日期范围内的实际工作日天数（考虑节假日）

    参数:
        start_date: 开始日期 (datetime.date 或 str 'YYYY-MM-DD')
        end_date: 结束日期 (datetime.date 或 str 'YYYY-MM-DD')
        workdays: 工作日列表 [1,2,3,4,5] (1=周一, 7=周日)
        holidays: 节假日列表 [datetime.date, ...] 或 [{'holidayDate': '2026-01-01', 'isWorkday': False}, ...] (可选)

    返回: 工作日天数
    """
    # 转换日期格式
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    # 处理节假日列表
    holiday_set = set()
    workday_set = set()  # 调休工作日

    if holidays:
        for item in holidays:
            if isinstance(item, dict):
                # 字典格式 {'holidayDate': '2026-01-01', 'isWorkday': False}
                h_date = isinstance(item['holidayDate'], str) and datetime.strptime(item['holidayDate'], '%Y-%m-%d').date() or item['holidayDate']
                if item.get('isWorkday', False):
                    workday_set.add(h_date)
                else:
                    holiday_set.add(h_date)
            elif isinstance(item, (date, datetime)):
                # 直接是日期对象
                holiday_set.add(item)

    # 计算工作日
    workdays_list = get_workdays_in_range(start_date, end_date, workdays, holiday_set)

    # 加上调休工作日
    for wd in workday_set:
        if start_date <= wd <= end_date and wd not in workdays_list:
            workdays_list.append(wd)

    # 排序
    workdays_list.sort()

    return len(workdays_list)
