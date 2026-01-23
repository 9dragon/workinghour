import openpyxl

# 读取Excel文件
file_path = 'D:/work/projects/workinghour/requirements/工时日志 V4.1-20260105111747.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

output = []
output.append('=' * 100)
output.append('工时日志 Excel 文件完整字段说明')
output.append('=' * 100)
output.append('')
output.append('文件名: 工时日志 V4.1-20260105111747.xlsx')
output.append(f'总列数: {ws.max_column}')
output.append(f'总行数: {ws.max_row}')
output.append('')

output.append('=' * 100)
output.append('一、表头结构说明')
output.append('=' * 100)
output.append('')
output.append('该 Excel 文件采用两层表头结构:')
output.append('- 第1行：工时大类（如【项目交付工时】、【产研项目工时】等）')
output.append('- 第2行：具体字段（如项目经理、项目名称、工作时长、加班时长等）')
output.append('')
output.append('注意：部分列在第1行没有标题，表示这些列属于基础信息列，不归类到任何工时大类。')
output.append('')

output.append('=' * 100)
output.append('二、字段分类说明')
output.append('=' * 100)
output.append('')

# 基础信息字段
output.append('1. 基础信息字段 (列1-11)')
output.append('-' * 100)
basic_fields = [
    (1, '序号', '记录序号'),
    (2, '数据id', '钉钉系统唯一标识ID'),
    (3, '姓名', '员工姓名'),
    (4, '开始时间', '工时周期开始日期'),
    (5, '结束时间', '工时周期结束日期'),
    (6, '工作摘要/本周复盘', '本周工作总结'),
    (7, '下周计划', '下周工作计划'),
    (8, '所需支持', '需要什么支持'),
    (9, '附件', '附件文件'),
    (10, '当前时间', '填写时间'),
    (11, '当前地点', '填写地点'),
]
for col, name, desc in basic_fields:
    output.append(f'  列{col:2d}: {name:20s} - {desc}')

output.append('')
output.append('2. 项目交付工时 (列12-17)')
output.append('-' * 100)
project_delivery_fields = [
    (12, '【项目交付工时】', '项目交付-项目经理', '项目负责人'),
    (13, '', '项目交付-项目名称', '项目名称'),
    (14, '', '项目交付-工作时长', '工作时长（小时）'),
    (15, '', '项目交付-加班时长', '加班时长（小时）'),
    (16, '', '项目交付-工作内容', '具体工作内容描述'),
    (17, '', '项目交付-备注', '备注信息'),
]
for col, cat, name, desc in project_delivery_fields:
    if cat:
        output.append(f'  列{col:2d}: {cat} - {name:20s} - {desc}')
    else:
        output.append(f'  列{col:2d}: {name:20s} - {desc}')

output.append('')
output.append('3. 产研项目工时 (列18-23)')
output.append('-' * 100)
product_fields = [
    (18, '【产研项目工时】', '产品-项目经理', '产品经理'),
    (19, '', '产品-项目名称', '产品/项目名称'),
    (20, '', '产品-工作时长', '工作时长（小时）'),
    (21, '', '产品-加班时长', '加班时长（小时）'),
    (22, '', '产品-研发工作内容', '研发工作内容描述'),
    (23, '', '产品-备注', '备注信息'),
]
for col, cat, name, desc in product_fields:
    if cat:
        output.append(f'  列{col:2d}: {cat} - {name:20s} - {desc}')
    else:
        output.append(f'  列{col:2d}: {name:20s} - {desc}')

output.append('')
output.append('4. 售前支持工时 (列24-29)')
output.append('-' * 100)
presales_fields = [
    (24, '【售前支持工时】', '售前-审批人', '售前审批人'),
    (25, '', '售前-项目名称', '售前项目名称'),
    (26, '', '售前-工作时长', '工作时长（小时）'),
    (27, '', '售前-加班时长', '加班时长（小时）'),
    (28, '', '售前-工作内容', '售前工作内容描述'),
    (29, '', '售前-备注', '备注信息'),
]
for col, cat, name, desc in presales_fields:
    if cat:
        output.append(f'  列{col:2d}: {cat} - {name:20s} - {desc}')
    else:
        output.append(f'  列{col:2d}: {name:20s} - {desc}')

output.append('')
output.append('5. 部门内务工时 (列30-35)')
output.append('-' * 100)
dept_fields = [
    (30, '【部门内务工时】', '部门', '部门名称'),
    (31, '', '部门-项目名称', '内务项目名称'),
    (32, '', '部门-工作时长', '工作时长（小时）'),
    (33, '', '部门-加班时长', '加班时长（小时）'),
    (34, '', '部门-内务工作内容', '内务工作内容描述'),
    (35, '', '部门-备注', '备注信息'),
]
for col, cat, name, desc in dept_fields:
    if cat:
        output.append(f'  列{col:2d}: {cat} - {name:20s} - {desc}')
    else:
        output.append(f'  列{col:2d}: {name:20s} - {desc}')

output.append('')
output.append('6. 请假记录 (列36-37)')
output.append('-' * 100)
leave_fields = [
    (36, '【请假记录】', '请假类别', '请假类型（如年假、事假等）'),
    (37, '', '请假时长', '请假时长（小时）'),
]
for col, cat, name, desc in leave_fields:
    if cat:
        output.append(f'  列{col:2d}: {cat} - {name:20s} - {desc}')
    else:
        output.append(f'  列{col:2d}: {name:20s} - {desc}')

output.append('')
output.append('7. 审批流程字段 (列38-49)')
output.append('-' * 100)
approval_fields = [
    (38, '审批编号', '审批系统编号'),
    (39, '创建时间', '审批单创建时间'),
    (40, '创建人', '审批单创建人'),
    (41, '当前负责人', '当前审批负责人'),
    (42, '审批结果', '审批结果（通过/驳回）'),
    (43, '审批状态', '审批状态（已完成/进行中）'),
    (44, '更新时间', '最后更新时间'),
    (45, '创建人部门', '创建人所属部门'),
    (46, '审批单标题', '审批单标题'),
    (47, '历史审批人', '历史审批人列表'),
    (48, '耗时(时:分:秒)', '审批耗时'),
    (49, '审批记录', '详细审批记录'),
]
for col, name, desc in approval_fields:
    output.append(f'  列{col:2d}: {name:20s} - {desc}')

output.append('')
output.append('=' * 100)
output.append('三、工时字段命名规则分析')
output.append('=' * 100)
output.append('')
output.append('1. 工时大类分类：')
output.append('   - 【项目交付工时】：项目交付类工作')
output.append('   - 【产研项目工时】：产品研发类工作')
output.append('   - 【售前支持工时】：售前支持类工作')
output.append('   - 【部门内务工时】：部门内部管理类工作')
output.append('   - 【请假记录】：请假类记录')
output.append('')
output.append('2. 每个工时大类包含的子字段：')
output.append('   - 项目经理/审批人/部门：负责人')
output.append('   - 项目名称：工作所属项目')
output.append('   - 工作时长：正常工作时长（小时）')
output.append('   - 加班时长：加班时长（小时）')
output.append('   - 工作内容/研发工作内容/内务工作内容：具体工作描述')
output.append('   - 备注：其他说明')
output.append('')
output.append('3. 字段命名模式：')
output.append('   - 格式：{工时大类}-{具体字段}')
output.append('   - 例如：项目交付-工作时长、产品-加班时长、售前-项目名称')
output.append('')
output.append('4. 工时列汇总（共9列）：')
output.append('   - 列14: 项目交付-工作时长')
output.append('   - 列15: 项目交付-加班时长')
output.append('   - 列20: 产品-工作时长')
output.append('   - 列21: 产品-加班时长')
output.append('   - 列26: 售前-工作时长')
output.append('   - 列27: 售前-加班时长')
output.append('   - 列32: 部门-工作时长')
output.append('   - 列33: 部门-加班时长')
output.append('   - 列37: 请假时长')
output.append('')

output.append('=' * 100)
output.append('四、数据导入注意事项')
output.append('=' * 100)
output.append('')
output.append('1. 审批状态筛选：')
output.append('   - 仅导入 审批结果="通过" 且 审批状态="已完成" 的记录')
output.append('   - 列42: 审批结果')
output.append('   - 列43: 审批状态')
output.append('')
output.append('2. 唯一性标识：')
output.append('   - 使用 序号 + 姓名 + 开始时间 + 项目名称 作为唯一性标识')
output.append('   - 注意：由于有多个项目名称列，需要处理多行合并的情况')
output.append('')
output.append('3. 工时数据处理：')
output.append('   - 每条记录可能包含多个工时类型的时长（项目、产品、售前、部门）')
output.append('   - 需要将不同工时类型拆分为独立的数据行')
output.append('   - 例如：一条记录同时有项目交付工时和产研项目工时，应拆分为2条数据')
output.append('')
output.append('4. 空值处理：')
output.append('   - 工时列（工作时长、加班时长）为空的，表示该类型无工时')
output.append('   - 项目名称为空的，该行工时不记录到数据库')
output.append('')

# 写入文件
with open('D:/work/projects/workinghour/excel_fields_detailed.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(output))

print('详细字段说明已保存到: D:/work/projects/workinghour/excel_fields_detailed.txt')
